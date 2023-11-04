import requests
from bs4 import BeautifulSoup
import openpyxl
import shodan


with open('shodan_api_key.txt', 'r') as archivo_api:
    clave_api_shodan = archivo_api.read().strip()


api = shodan.Shodan(clave_api_shodan)

url = 'https://www.uanl.mx/?'
res = requests.get(url)
soup = BeautifulSoup(res.text, 'lxml')

links = []
for link in soup.find_all('a'):
    href = link.get('href')
    if href:
        links.append(href)

workbook_links = openpyxl.Workbook()
sheet_links = workbook_links.active
sheet_links.title = 'Enlaces'

for i, link in enumerate(links, 1):
    sheet_links.cell(row=i, column=1, value=link)


workbook_shodan = openpyxl.Workbook()
sheet_shodan = workbook_shodan.active
sheet_shodan.title = 'Resultados Shodan'


direccion_ip = '148.234.5.222'  
try:
    info_ips = api.search(direccion_ip)
    print("Resultados de Shodan para la IP:", direccion_ip)
  
    for index, result in enumerate(info_ips['matches'][:5], 1):
        sheet_shodan.cell(row=index, column=1, value=f"IP: {result['ip_str']}")
        sheet_shodan.cell(row=index, column=2, value=f"Org: {result.get('org', 'N/A')}")
        sheet_shodan.cell(row=index, column=3, value=f"OS: {result.get('os', 'N/A')}")
        sheet_shodan.cell(row=index, column=4, value=f"Puertos: {', '.join(map(str, result.get('puertos', [])))}")

except shodan.APIError as e:
    print(f"Error: {e}")

workbook_links.save('links.xlsx')
workbook_shodan.save('shodan_results.xlsx')

