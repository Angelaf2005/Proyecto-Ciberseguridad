import requests
from bs4 import BeautifulSoup
import openpyxl
import shodan
import logging
import zipfile
def API_Shodan(direccion_ip):
    logging.basicConfig(filename='app.log', level=logging.INFO)
    with open("API_KEY\shodan_api_key.txt", 'r') as archivo_api:
        clave_api_shodan = archivo_api.read().strip()
    logging.info(clave_api_shodan)
    api = shodan.Shodan(clave_api_shodan)
    workbook_shodan = openpyxl.Workbook()
    sheet_shodan = workbook_shodan.active
    sheet_shodan.title = 'Resultados Shodan'
    info_ips = api.search(direccion_ip)
    print("Resultados de Shodan para la IP:", direccion_ip)
    for index, result in enumerate(info_ips['matches'][:5], 1):
        sheet_shodan.cell(row=index, column=1, value=f"IP: {result['ip_str']}")
        sheet_shodan.cell(row=index, column=2, value=f"Org: {result.get('org', 'N/A')}")
        sheet_shodan.cell(row=index, column=3, value=f"OS: {result.get('os', 'N/A')}")
    workbook_shodan.save('archivos/shodan_results.xlsx')
def API_hackertarget(ip):
    try:
        r = requests.get("https://api.hackertarget.com/ipgeo/?q="+ip)
        if r.status_code == 200:
            with open("geolocalization.txt","w") as archivo:
                archivo.write(r.text)
        with zipfile.ZipFile("./pass/archivos.zip","a") as archivo:
                archivo.write("./archivos/geolocalization.txt")
    except:
         pass